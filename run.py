from gd import app,mdb
from flask import render_template,make_response

import datetime

import json
@app.route('/test')
def test():
    for c in mdb.clivages.find():
        tot = c['g']+c['d']
        if tot>600:
            mdb.clivages.update({'i':c['i']},{'$set':{'g':0,'d':0}})
            print c
    #mdb.logs.remove()
    return "ok"
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook
    wb = Workbook()
    ws = wb.active
    for i,cliv in enumerate(mdb.clivages.find({})):
        ws.cell(row=i+1,column=1).value = cliv['i']
        ws.cell(row=i+1,column=2).value = cliv['n']
    resp = make_response(save_virtual_workbook(wb))
    send_file(dl_fd,
                 mimetype='application/pdf',
                 as_attachment=True,
                 attachment_filename=filename)
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = '"attachment; filename="clivage%s.xlsx"' % datetime.datetime.now().strftime('%Y-%m-%d')
    return resp



if __name__ == "__main__":
    app.run()
